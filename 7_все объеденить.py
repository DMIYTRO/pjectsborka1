import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from copy import copy
import shutil

# Путь к папке с исходными файлами
folder_path = r'C:\Users\Dmitry\Downloads\сборки_xls'

# Путь к папке для перемещения файлов
done_folder_path = os.path.join(folder_path, 'done')

# Создаем папку "done", если она не существует
if not os.path.exists(done_folder_path):
    os.makedirs(done_folder_path)

# Список колонок, для которых необходимо установить максимальный размер
columns_to_resize = ['B', 'C']

# Создаем новый файл для объединенных данных
merged_workbook = Workbook()
merged_sheet = merged_workbook.active

# Устанавливаем цвет для выделения строк
highlight_color = "FFFF00"  # Желтый цвет

# Проходим по всем файлам в указанной папке
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(folder_path, filename)
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Копируем данные из текущего файла в объединенный файл
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            merged_sheet.append(row)

            # Проверяем значение в столбце "Заказ от" (здесь предполагается, что столбец "Заказ от" находится в колонке B)
            if row[1] == "Заказ от":
                # Изменяем цвет всей строки на заданный цвет
                for cell in merged_sheet[-1]:
                    cell.fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type="solid")

            # Копируем стили ячеек из второго скрипта
            for col, cell_value in enumerate(row, start=1):
                src_cell = sheet.cell(row=idx + 1, column=col)
                dest_cell = merged_sheet.cell(row=merged_sheet.max_row, column=col)
                if src_cell.has_style:
                    dest_cell.font = copy(src_cell.font)
                    dest_cell.border = copy(src_cell.border)
                    dest_cell.fill = copy(src_cell.fill)
                    dest_cell.number_format = copy(src_cell.number_format)
                    dest_cell.protection = copy(src_cell.protection)
                    dest_cell.alignment = copy(src_cell.alignment)

        # Перемещаем исходный файл в папку "done"
        shutil.move(file_path, os.path.join(done_folder_path, filename))

# Устанавливаем максимальный размер (60 символов) для указанных колонок в объединенном файле
for column_letter in columns_to_resize:
    merged_sheet.column_dimensions[column_letter].width = 60

# Сохраняем объединенный файл
merged_file_path = os.path.join(folder_path, 'merged.xlsx')
merged_workbook.save(merged_file_path)

print('Объединение файлов завер шено')

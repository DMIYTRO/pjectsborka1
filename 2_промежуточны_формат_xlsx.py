import pandas as pd
import glob
from openpyxl import load_workbook

folder_path = r'C:\Users\Dmitry\Downloads\сборки_xls'
file_list = glob.glob(folder_path + '\\*.xlsx')

dfs = []
saved_values = []  # Инициализация переменной saved_values

# Проход по каждому файлу и получение значений saved_values
for file in file_list:
    df = pd.read_excel(file)
    dfs.append(df)

    # Получение значений первых двух строк первого столбца
    column_name = df.columns[0]  # Имя первого столбца
    values = df.iloc[:2, 0].tolist()  # Значения первых двух строк первого столбца, сохраненные в список
    saved_values.extend(values)

    # Сохранение column_name и values в исходном файле
    book = load_workbook(file)
    writer = pd.ExcelWriter(file, engine='openpyxl')
    writer.book = book

    info_sheet = book.active
    # info_sheet.cell(row=len(df) + 2, column=1, value='column_name:')
    # info_sheet.cell(row=len(df) + 2, column=2, value=column_name)
    # info_sheet.cell(row=len(df) + 3, column=1, value='СБОРКА:')
    for i, value in enumerate(values):
        info_sheet.cell(row=len(df) + 2 + i, column=2, value=value)

    writer.close()

print(f"Saved values:")
print(saved_values)

# Проход по каждому файлу и удаление указанных колонок
for file in file_list:
    df = pd.read_excel(file, header=3)  # Указываем номер строки заголовка (считая с 0)
    columns_to_delete = ['продукт', 'пост', 'п+пп', 'пп', 'комментарий', 'препресс']
    df = df.drop(columns=columns_to_delete)

    df.to_excel(file, index=False)  # Перезаписываем файл с обновленными данными

    print(f"Удалены колонки в файле: {file}")

# Объединение загруженных файлов в датафрейм
if dfs:
    df_combined = pd.concat(dfs, ignore_index=True)
    column_name = df_combined.columns[0]  # Имя первого столбца

    print(f"Values from column '{column_name}':")
    print(saved_values)
else:
    print("No data to concatenate.")

# Сохранение значений в переменную
values = saved_values

# Пример использования сохраненных значений
print("Saved values:")
print(values)
